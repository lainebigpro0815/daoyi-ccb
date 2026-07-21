from datetime import date
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.project import Project, ProjectPhase, ProjectTask

router = APIRouter(prefix="/api/projects/{project_id}/export", tags=["export"])


@router.get("/excel")
def export_excel(project_id: int, db: Session = Depends(get_db)):
    """导出项目计划为 Excel 文件"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=400, detail="需要安装 openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目计划"

    # === 样式定义 ===
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1D2129", end_color="1D2129", fill_type="solid")
    phase_font = Font(name="微软雅黑", bold=True, size=10, color="1D2129")
    phase_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    task_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="center")

    # === 表头 ===
    headers = ["任务编号", "任务名称", "负责人", "计划开始", "计划结束", "进度", "状态", "输出物"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 列宽
    widths = [12, 35, 12, 13, 13, 8, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # === 项目信息行 ===
    ws.merge_cells("A2:H2")
    info_cell = ws.cell(row=2, column=1,
                        value=f"项目：{project.name} | 客户：{project.customer_name} | 时间：{project.start_date} → {project.planned_end_date or ''}")
    info_cell.font = Font(name="微软雅黑", size=10, color="666666")
    ws.row_dimensions[2].height = 28

    # === 阶段 + 任务 ===
    row = 3
    phases = db.query(ProjectPhase).filter(
        ProjectPhase.project_id == project_id
    ).order_by(ProjectPhase.sort_order).all()

    for phase in phases:
        tasks = db.query(ProjectTask).filter(
            ProjectTask.project_phase_id == phase.id
        ).order_by(ProjectTask.sort_order).all()

        # 阶段行
        ws.merge_cells(f"A{row}:H{row}")
        phase_cell = ws.cell(row=row, column=1,
                             value=f"阶段{phase.phase_number}：{phase.name} ({phase.planned_start} → {phase.planned_end})")
        phase_cell.font = phase_font
        phase_cell.fill = phase_fill
        phase_cell.border = thin_border
        ws.row_dimensions[row].height = 26
        row += 1

        for t in tasks:
            status_map = {"pending": "待开始", "in_progress": "进行中",
                          "completed": "已完成", "blocked": "阻塞"}
            vals = [
                t.task_number, t.name,
                t.assignee or "",
                str(t.planned_start or ""), str(t.planned_end or ""),
                f"{t.progress}%",
                status_map.get(t.status, t.status),
                t.deliverable or "",
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = task_font
                cell.border = thin_border
                cell.alignment = wrap_align
            ws.row_dimensions[row].height = 22
            row += 1

    # === 统计行 ===
    row += 1
    total_tasks = sum(len(db.query(ProjectTask).filter(
        ProjectTask.project_phase_id == p.id).all()) for p in phases)
    completed_tasks = sum(
        sum(1 for t in db.query(ProjectTask).filter(
            ProjectTask.project_phase_id == p.id).all() if t.status == "completed")
        for p in phases
    )
    ws.merge_cells(f"A{row}:H{row}")
    stat = ws.cell(row=row, column=1,
                   value=f"统计：共 {total_tasks} 个任务，已完成 {completed_tasks} 个，进度 {round(completed_tasks/total_tasks*100 if total_tasks else 0)}%")
    stat.font = Font(name="微软雅黑", size=10, bold=True, color="1D2129")

    # 保存
    export_dir = Path(__file__).parent.parent.parent / "data" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    filepath = export_dir / f"project_{project_id}_{date.today().isoformat()}.xlsx"
    wb.save(str(filepath))

    return FileResponse(
        path=str(filepath),
        filename=f"{project.name}_项目计划.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
